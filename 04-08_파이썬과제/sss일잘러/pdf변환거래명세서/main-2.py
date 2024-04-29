from openpyxl import load_workbook
from datetime import date

견적서_파일경로 = r"pdf변환거래명세서\견적서_샘플.xlsx"

견적서_wb = load_workbook(견적서_파일경로, data_only=True)
견적서_ws = 견적서_wb.active

# 하나의 셀 읽기
견적받는자 = 견적서_ws['A4'].value

# 여러 개의 셀 반복하여 읽기
품목명_list = []
수량_list = []
단가_list = []
금액_list = []

for data in 견적서_ws['C13':'X24']:
    for cell in data:
        if cell.column == ord('C') - 64:
            if cell.value is not None:
                품목명_list.append(cell.value)
        if cell.column == ord('R') - 64:
            if cell.value is not None:
                수량_list.append(cell.value)
        if cell.column == ord('T') - 64:
            if cell.value is not None:
                단가_list.append(cell.value)
        if cell.column == ord('X') - 64:
            if cell.value is not None:
                금액_list.append(cell.value)
print("픔목명: ",품목명_list)
print("수량: ",수량_list)
print("단가: ", 단가_list)
print("금액", 금액_list)

# 오늘 날짜 구하기
today = date.today()
# 거래명세표 엑셀 파일에 오늘 날짜 입력하기
거래명세표_파일경로 = r"pdf변환거래명세서\거래명세표_샘플.xlsx"
거래명세표_wb = load_workbook(거래명세표_파일경로, data_only=False)
거래명세표_ws = 거래명세표_wb.active

거래명세표_ws['C4'].value = today.year
거래명세표_ws['E4'].value = today.month
거래명세표_ws['G4'].value = today.day

거래명세표_ws['C6'].value = 견적받는자

for i in range(len(품목명_list)):
    거래명세표_ws.cell(row=i + 10, column=ord('B') - 64).value = 품목명_list[i]
    거래명세표_ws.cell(row=i + 10, column=ord('G') - 64).value = 수량_list[i]
    거래명세표_ws.cell(row=i + 10, column=ord('I') - 64).value = 단가_list[i]

# 수정된 엑셀 파일 저장
거래명세표_wb.save("pdf변환거래명세서\\" + "거래명세표_" + 견적받는자 + ".xlsx")
